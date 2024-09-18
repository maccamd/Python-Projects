import openpyxl as xl
from openpyxl.utils import get_column_letter
import tkinter as tk
from tkinter import Image, filedialog
from tkinter import messagebox
import matplotlib.cbook as cbook
import matplotlib as image
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages
from matplotlib.cbook import get_sample_data
from matplotlib.offsetbox import  OffsetImage
from PIL import Image

#Driver Trace Analysis Tool
#Version 1.74 of code

#----tkinter----#

root = tk.Tk()
root.withdraw()

#----Functions----#
def wltc (time, target_speed, actual_speed, upper_boundary, lower_boundary, filename, filecut, driver, sheet, errorcnt1, errortime1, violationcnt1, errorcnt2, errortime2, violationcnt2, violationtime2, iwr2, errorcnt3, errortime3, violationcnt3, violationtime3, iwr3, errorcnt4, errortime4, violationcnt4, violationtime4, iwr4):
    #----Slices whole trace data into individual phase data----#
    phase1_time = time[0:5890]
    phase1_target_speed = target_speed[0:5890]
    phase1_actual_speed = actual_speed[0:5890]
    phase1_upper_boundary = upper_boundary[0:5890]
    phase1_lower_boundary = lower_boundary[0:5890]

    phase2_time = time[5891:10220]
    phase2_target_speed = target_speed[5891:10220]
    phase2_actual_speed = actual_speed[5891:10220] 
    phase2_upper_boundary = upper_boundary[5891:10220]
    phase2_lower_boundary = lower_boundary[5891:10220]

    phase3_time = time[10221:14770]
    phase3_target_speed = target_speed[10221:14770]
    phase3_actual_speed = actual_speed[10221:14770]
    phase3_upper_boundary = upper_boundary[10221:14770]
    phase3_lower_boundary = lower_boundary[10221:14770]

    phase4_time = time[14771:]
    phase4_target_speed = target_speed[14771:]
    phase4_actual_speed = actual_speed[14771:]
    phase4_upper_boundary = upper_boundary[14771:]
    phase4_lower_boundary = lower_boundary[14771:]

    #----matplotlib section----#
    #Figure 1
    fig, axs1 = plt.subplots(2, figsize=(11.69,8.27))
    plt.subplots_adjust(left=0.05, bottom=0.06, right=0.97, top=0.93, wspace=0.2, hspace=0.23)
    fig.suptitle(filename, x=0.1, y=0.95, horizontalalignment='left', verticalalignment='top')
    newax = fig.add_axes([0.84, 0.881, 0.124, 0.115], anchor='NE',)
    newax.imshow(logo)
    newax.axis('off')    
    axs1[0].set_title('Phase 1')
    axs1[0].set_xlabel('Time (s)')
    axs1[0].set_ylabel('Speed (Km/h)')
    axs1[0].text(0, 66, 'Driver: ' + driver)
    axs1[0].text(0, 63, 'Error Count: ' + errorcnt1 + ',  Error Time: ' + errortime1)
    axs1[0].text(0, 60.5, 'Violation Count: ' + violationcnt1 + ',  Violation Time: ' + violationtime1 + ', IWR Phase 1: ' + iwr1 + '%')
    axs1[0].text(410, 63, 'IWR Test: ' + iwrtest + '%')
    axs1[0].plot(phase1_time,phase1_target_speed, label="Target Speed", linewidth= 0.75) # line width unreadable below 0.5
    axs1[0].plot(phase1_time,phase1_actual_speed, label="Actual Speed",linewidth= 0.75)
    axs1[0].plot(phase1_time,phase1_upper_boundary, linestyle= '--',linewidth= 0.75, color= 'grey')
    axs1[0].plot(phase1_time,phase1_lower_boundary, linestyle= '--',linewidth= 0.75, color= 'grey')
    axs1[0].legend()
    axs1[0].grid(linestyle = '--', linewidth= 0.5)
    axs1[0].margins(0)
    axs1[0].set_ylim([-0.3,60])

    axs1[1].set_title('Phase 2')
    axs1[1].set_xlabel('Time (s)')
    axs1[1].set_ylabel('Speed (Km/h)')
    axs1[1].text(589, 84, 'Error Count: ' + errorcnt2 + ',  Error Time: ' + errortime2)
    axs1[1].text(589, 80.5, 'Violation Count: ' + violationcnt2 + ',  Violation Time: ' + violationtime2 + ', IWR Phase 2: ' + iwr2 + '%')
    axs1[1].plot(phase2_time,phase2_target_speed, label="Target Speed", linewidth= 0.75,)
    axs1[1].plot(phase2_time,phase2_actual_speed, label="Actual Speed", linewidth= 0.75)
    axs1[1].plot(phase2_time,phase2_upper_boundary, linestyle= '--',linewidth= 0.75, color= 'grey')
    axs1[1].plot(phase2_time,phase2_lower_boundary, linestyle= '--',linewidth= 0.75, color= 'grey')
    axs1[1].legend()
    axs1[1].grid(linestyle = '--', linewidth= 0.5)
    axs1[1].margins(0)
    axs1[1].set_ylim([-0.3,80])

    #Figure 2
    fig2, axs2 = plt.subplots(2, figsize=(11.96,8.27))
    plt.subplots_adjust(left=0.05, bottom=0.06, right=0.97, top=0.93, wspace=0.2, hspace=0.23)
    fig.suptitle(filename)
    newax = fig2.add_axes([0.84, 0.881, 0.124, 0.115], anchor='NE',)
    newax.imshow(logo)
    newax.axis('off')    
    axs2[0].set_title('Phase 3')
    axs2[0].set_xlabel('Time (s)')
    axs2[0].set_ylabel('Speed (Km/h)')
    axs2[0].text(1022, 106, 'Error Count: ' + errorcnt3 + ',  Error Time: ' + errortime3)
    axs2[0].text(1022, 101.5, 'Violation Count: ' + violationcnt3 + ',  Violation Time: ' + violationtime3 + ', IWR Phase 3: ' + iwr3 + '%')
    axs2[0].plot(phase3_time,phase3_target_speed, label="Target Speed", linewidth= 0.75)
    axs2[0].plot(phase3_time,phase3_actual_speed, label="Actual Speed", linewidth= 0.75)
    axs2[0].plot(phase3_time,phase3_upper_boundary, linestyle= '--',linewidth= 0.75, color= 'grey')
    axs2[0].plot(phase3_time,phase3_lower_boundary, linestyle= '--',linewidth= 0.75, color= 'grey')
    axs2[0].legend()
    axs2[0].grid(linestyle = '--', linewidth= 0.5)
    axs2[0].margins(0)
    axs2[0].set_xlim([1022,1477])
    axs2[0].set_ylim([-0.3,101])

    axs2[1].set_title('Phase 4')
    axs2[1].set_xlabel('Time (s)')
    axs2[1].set_ylabel('Speed (Km/h)')
    axs2[1].text(1477, 142, 'Error Count: ' + errorcnt4 + ',  Error Time: ' + errortime4)
    axs2[1].text(1477, 136.8, 'Violation Count: ' + violationcnt4 + ',  Violation Time: ' + violationtime4 + ', IWR Phase 4: ' + iwr4 + '%')
    axs2[1].plot(phase4_time,phase4_target_speed, label="Target Speed", linewidth= 0.75)
    axs2[1].plot(phase4_time,phase4_actual_speed, label="Actual Speed", linewidth= 0.75)
    axs2[1].plot(phase4_time,phase4_upper_boundary, linestyle= '--',linewidth= 0.75, color= 'grey')
    axs2[1].plot(phase4_time,phase4_lower_boundary, linestyle= '--',linewidth= 0.75, color= 'grey')
    axs2[1].legend()
    axs2[1].grid(linestyle = '--', linewidth= 0.5)
    axs2[1].margins(0)
    axs2[1].set_xlim([1477,1820])
    axs2[1].set_ylim([-0.3,135])

    filename = PdfPages(filecut + ".pdf")
    filename.savefig(fig)
    filename.savefig(fig2)
    filename.close()
    
    return
def nedc (time, target_speed, actual_speed, upper_boundary, lower_boundary, filename, filecut, driver, sheet, errorcnt1, errortime1, violationcnt1, errorcnt2, errortime2, violationcnt2, violationtime2):
    phase1_time = time[0:7799]
    phase1_target_speed = target_speed[0:7799]
    phase1_actual_speed = actual_speed[0:7799]
    phase1_upper_boundary = upper_boundary[0:7799]
    phase1_lower_boundary = lower_boundary[0:7799]

    phase2_time = time[7800:12008]
    phase2_target_speed = target_speed[7800:12008]
    phase2_actual_speed = actual_speed[7800:12008] 
    phase2_upper_boundary = upper_boundary[7800:12008]
    phase2_lower_boundary = lower_boundary[7800:12008]

    fig, axs1 = plt.subplots(2, figsize=(11.69,8.27))
    plt.subplots_adjust(left=0.05, bottom=0.06, right=0.97, top=0.93, wspace=0.2, hspace=0.23)
    fig.suptitle(filename)
    newax = fig.add_axes([0.84, 0.881, 0.124, 0.115], anchor='NE',)
    newax.imshow(logo)
    newax.axis('off')
    axs1[0].set_title('Phase 1')
    axs1[0].set_xlabel('Time (s)')
    axs1[0].set_ylabel('Speed (Km/h)')
    axs1[0].text(0, 60.5, 'Driver: ' + driver)
    axs1[0].text(0, 58, 'Error Count: ' + errorcnt1 + ',  Error Time: ' + errortime1)
    axs1[0].text(0, 55.5, 'Violation Count: ' + violationcnt1 + ',  Violation Time: ' + violationtime1)
    axs1[0].plot(phase1_time,phase1_target_speed, label="Target Speed", linewidth= 0.75) # line width unreadable below 0.5
    axs1[0].plot(phase1_time,phase1_actual_speed, label="Actual Speed",linewidth= 0.75)
    axs1[0].plot(phase1_time,phase1_upper_boundary, linestyle= '--',linewidth= 0.75, color= 'grey')
    axs1[0].plot(phase1_time,phase1_lower_boundary, linestyle= '--',linewidth= 0.75, color= 'grey')
    axs1[0].legend()
    axs1[0].grid(linestyle = '--', linewidth= 0.5)
    axs1[0].margins(0)
    axs1[0].set_ylim([-0.3,55])

    axs1[1].set_title('Phase 2')
    axs1[1].set_xlabel('Time (s)')
    axs1[1].set_ylabel('Speed (Km/h)')
    axs1[1].text(780, 130.5, 'Error Count: ' + errorcnt2 + ',  Error Time: ' + errortime2)
    axs1[1].text(780, 125.5, 'Violation Count: ' + violationcnt2 + ',  Violation Time: ' + violationtime2)
    axs1[1].plot(phase2_time,phase2_target_speed, label="Target Speed", linewidth= 0.75,)
    axs1[1].plot(phase2_time,phase2_actual_speed, label="Actual Speed", linewidth= 0.75)
    axs1[1].plot(phase2_time,phase2_upper_boundary, linestyle= '--',linewidth= 0.75, color= 'grey')
    axs1[1].plot(phase2_time,phase2_lower_boundary, linestyle= '--',linewidth= 0.75, color= 'grey')
    axs1[1].legend()
    axs1[1].grid(linestyle = '--', linewidth= 0.5)
    axs1[1].margins(0)
    axs1[1].set_ylim([-0.3,125])

    filename = PdfPages(filecut + ".pdf")
    filename.savefig(fig)
    filename.close()
    return
def us06x1 (time, target_speed, actual_speed, upper_boundary, lower_boundary, filename, filecut, driver, sheet, errorcnt1, errortime1, violationcnt1):
    phase1_time = time[0:6208]
    phase1_target_speed = target_speed[0:6208]
    phase1_actual_speed = actual_speed[0:6208]
    phase1_upper_boundary = upper_boundary[0:6208]
    phase1_lower_boundary = lower_boundary[0:6208]

    fig, axs1 = plt.subplots(1, figsize=(11.69,8.27))
    plt.subplots_adjust(left=0.05, bottom=0.06, right=0.97, top=0.93, wspace=0.2, hspace=0.23)
    fig.suptitle(filename)
    newax = fig.add_axes([0.84, 0.881, 0.124, 0.115], anchor='NE',)
    newax.imshow(logo)
    newax.axis('off')
    axs1.set_title('US06 Cycle')
    axs1.set_xlabel('Time (s)')
    axs1.set_ylabel('Speed (Km/h)')
    axs1.text(0, 147, 'Driver: ' + driver)
    axs1.text(0, 144, 'Error Count: ' + errorcnt1 + ',  Error Time: ' + errortime1)
    axs1.text(0, 141, 'Violation Count: ' + violationcnt1 + ',  Violation Time: ' + violationtime1)
    axs1.plot(phase1_time,phase1_target_speed, label="Target Speed", linewidth= 0.75) # line width unreadable below 0.5
    axs1.plot(phase1_time,phase1_actual_speed, label="Actual Speed",linewidth= 0.75)
    axs1.plot(phase1_time,phase1_upper_boundary, linestyle= '--',linewidth= 0.75, color= 'grey')
    axs1.plot(phase1_time,phase1_lower_boundary, linestyle= '--',linewidth= 0.75, color= 'grey')
    axs1.legend()
    axs1.grid(linestyle = '--', linewidth= 0.5)
    axs1.margins(0)
    axs1.set_ylim([-0.3,140])

    filename = PdfPages(filecut + ".pdf")
    filename.savefig(fig)
    filename.close()
    return
def hwfetx1 (time, target_speed, actual_speed, upper_boundary, lower_boundary, filename, filecut, driver, sheet, errorcnt1, errortime1, violationcnt1):
    phase1_time = time[0:7651]
    phase1_target_speed = target_speed[0:7651]
    phase1_actual_speed = actual_speed[0:7651]
    phase1_upper_boundary = upper_boundary[0:7651]
    phase1_lower_boundary = lower_boundary[0:7651]

    fig, axs1 = plt.subplots(1, figsize=(11.69,8.27))
    plt.subplots_adjust(left=0.05, bottom=0.06, right=0.97, top=0.93, wspace=0.2, hspace=0.23)
    fig.suptitle(filename)
    newax = fig.add_axes([0.84, 0.881, 0.124, 0.115], anchor='NE',)
    newax.imshow(logo)
    newax.axis('off')
    axs1.set_title('HWFET Cycle')
    axs1.set_xlabel('Time (s)')
    axs1.set_ylabel('Speed (Km/h)')
    axs1.text(0, 105, 'Driver: ' + driver)
    axs1.text(0, 103, 'Error Count: ' + errorcnt1 + ',  Error Time: ' + errortime1)
    axs1.text(0, 101, 'Violation Count: ' + violationcnt1 + ',  Violation Time: ' + violationtime1)
    axs1.plot(phase1_time,phase1_target_speed, label="Target Speed", linewidth= 0.75) # line width unreadable below 0.5
    axs1.plot(phase1_time,phase1_actual_speed, label="Actual Speed",linewidth= 0.75)
    axs1.plot(phase1_time,phase1_upper_boundary, linestyle= '--',linewidth= 0.75, color= 'grey')
    axs1.plot(phase1_time,phase1_lower_boundary, linestyle= '--',linewidth= 0.75, color= 'grey')
    axs1.legend()
    axs1.grid(linestyle = '--', linewidth= 0.5)
    axs1.margins(0)
    axs1.set_ylim([-0.3,100])

    filename = PdfPages(filecut + ".pdf")
    filename.savefig(fig)
    filename.close()
    return
def sc03x1 (time, target_speed, actual_speed, upper_boundary, lower_boundary, filename, filecut, driver, sheet, errorcnt1, errortime1, violationcnt1):
    phase1_time = time[0:6000]
    phase1_target_speed = target_speed[0:6000]
    phase1_actual_speed = actual_speed[0:6000]
    phase1_upper_boundary = upper_boundary[0:6000]
    phase1_lower_boundary = lower_boundary[0:6000]
    
    fig, axs1 = plt.subplots(1, figsize=(11.69,8.27))
    plt.subplots_adjust(left=0.05, bottom=0.06, right=0.97, top=0.93, wspace=0.2, hspace=0.23)
    fig.suptitle(filename)
    newax = fig.add_axes([0.84, 0.881, 0.124, 0.115], anchor='NE',)
    newax.imshow(logo)
    newax.axis('off')
    axs1.set_title('SC03 Cycle')
    axs1.set_xlabel('Time (s)')
    axs1.set_ylabel('Speed (Km/h)')
    axs1.text(0, 97, 'Driver: ' + driver)
    axs1.text(0, 95, 'Error Count: ' + errorcnt1 + ',  Error Time: ' + errortime1)
    axs1.text(0, 93, 'Violation Count: ' + violationcnt1 + ',  Violation Time: ' + violationtime1)
    axs1.plot(phase1_time,phase1_target_speed, label="Target Speed", linewidth= 0.75) # line width unreadable below 0.5
    axs1.plot(phase1_time,phase1_actual_speed, label="Actual Speed",linewidth= 0.75)
    axs1.plot(phase1_time,phase1_upper_boundary, linestyle= '--',linewidth= 0.75, color= 'grey')
    axs1.plot(phase1_time,phase1_lower_boundary, linestyle= '--',linewidth= 0.75, color= 'grey')
    axs1.legend()
    axs1.grid(linestyle = '--', linewidth= 0.5)
    axs1.margins(0)
    axs1.set_ylim([-0.3,92])
    filename = PdfPages(filecut + ".pdf")
    filename.savefig(fig)
    filename.close()
    return
#----Open File----# 

file_path = filedialog.askopenfilename() #tkinter method
path = open(file_path, "rb") #rb is reading in binary
print(file_path)

#----Opens Excel sheet----#
wb = xl.load_workbook(path)
sheet = wb['ContinuousData']            #wb.active is sheet the file opens on

tarspeed = "" #Looking for 'DA Schedule Speed' column index
actspeed = "" #Looking for 'DA Actual Speed' column index
uppertol = "" #Looking for 'UpperTolerance' column index
lowertol = "" #Looking for 'LowerTolerance' column index

#logo = plt.imread('1200px-MAHLE_PTL_4c.png') N:/EDC Schedules/Drive Trace Analysis Tool V1.72/Current version/Drive Trace Analysis Tool V1.72/
logo = Image.open('MPT Logo.png')

#----Loop to get data out of sheet----#

for row_cells in sheet.iter_rows(min_row=1, max_row=1):
    for cell in row_cells:
        if 'DA Schedule Speed' in cell.value or 'ScheduledSpeed' in cell.value:
            tarspeed = get_column_letter(cell.column)

for row_cells in sheet.iter_rows(min_row=1, max_row=1):
    for cell in row_cells:
        if 'DA Actual Speed' in cell.value or 'SpeedFeedback' in cell.value:
            actspeed = get_column_letter(cell.column)

for row_cells in sheet.iter_rows(min_row=1, max_row=1):
    for cell in row_cells:
        if 'UpperTolerance' in cell.value:
            uppertol = get_column_letter(cell.column)

for row_cells in sheet.iter_rows(min_row=1, max_row=1):
    for cell in row_cells:
        if 'LowerTolerance' in cell.value:
            lowertol = get_column_letter(cell.column)

time = [sheet['B' + str(row)].value for row in range(3, sheet.max_row + 1)] #PlotTime
target_speed = [sheet[tarspeed + str(row)].value for row in range(3, sheet.max_row + 1)] #DA Schedule Speed
actual_speed = [sheet[actspeed + str(row)].value for row in range(3, sheet.max_row + 1)] #DA Actual Speed
upper_boundary = [sheet[uppertol + str(row)].value for row in range(3, sheet.max_row + 1)] #UpperTolerance
lower_boundary = [sheet[lowertol + str(row)].value for row in range(3, sheet.max_row + 1)] #LowerTolerance

sheet = wb['Summary']
filename = sheet['B1'].value
filecut = filename[:-21]
driver = sheet['AD3'].value
cycle = sheet['D2'].value
iwrtest = str(round(sheet['T76'].value, 2))

sheet = wb['Phase1']
errorcnt1 = str(sheet['D16'].value)
errortime1 = str(round(sheet['H16'].value, 3))
violationcnt1 = str(sheet['L16'].value)
iwr1 = str(round(sheet['T14'].value, 2))
violationtime1 = str(round(sheet['T16'].value, 3))

#----Conditions to choose which function to run----#
if cycle == 'WLTC Class3b TYPE 1' or cycle == 'WLTC Class3b TYPE 1 Development':

    sheet = wb['Phase2']
    errorcnt2 = str(sheet['D16'].value)
    errortime2 = str(round(sheet['H16'].value, 3))
    violationcnt2 = str(sheet['L16'].value)
    iwr2 = str(round(sheet['T14'].value, 2))
    violationtime2 = str(round(sheet['T16'].value, 3))

    sheet = wb['Phase3']
    errorcnt3 = str(sheet['D16'].value)
    errortime3 = str(round(sheet['H16'].value, 3))
    violationcnt3 = str(sheet['L16'].value)
    iwr3 = str(round(sheet['T14'].value, 2))
    violationtime3 = str(round(sheet['T16'].value, 3))

    sheet = wb['Phase4']
    errorcnt4 = str(sheet['D16'].value)
    errortime4 = str(round(sheet['H16'].value, 3))
    violationcnt4 = str(sheet['L16'].value)
    iwr4 = str(round(sheet['T14'].value, 2))
    violationtime4 = str(round(sheet['T16'].value, 3))
    wltc(time, target_speed, actual_speed, upper_boundary, lower_boundary, filename, filecut, driver, sheet, errorcnt1, errortime1, violationcnt1,  errorcnt2, errortime2, violationcnt2, violationtime2, iwr2, errorcnt3, errortime3, violationcnt3, violationtime3, iwr3, errorcnt4, errortime4, violationcnt4, violationtime4, iwr4)
elif cycle == 'NEDC TYPE 1' or cycle == 'MVEG B MT':
    sheet = wb['Phase2']
    errorcnt2 = str(sheet['D16'].value)
    errortime2 = str(round(sheet['H16'].value, 3))
    violationcnt2 = str(sheet['L16'].value)
    violationtime2 = str(round(sheet['T16'].value, 3))
    nedc(time, target_speed, actual_speed, upper_boundary, lower_boundary, filename, filecut, driver, sheet, errorcnt1, errortime1, violationcnt1, errorcnt2, errortime2, violationcnt2, violationtime2)
elif cycle == 'US06':
    us06x1(time, target_speed, actual_speed, upper_boundary, lower_boundary, filename, filecut, driver, sheet, errorcnt1, errortime1, violationcnt1)
elif cycle == 'HWFET':
     hwfetx1(time, target_speed, actual_speed, upper_boundary, lower_boundary, filename, filecut, driver, sheet, errorcnt1, errortime1, violationcnt1)
elif cycle == 'SC03':
     sc03x1(time, target_speed, actual_speed, upper_boundary, lower_boundary, filename, filecut, driver, sheet, errorcnt1, errortime1, violationcnt1)

messagebox.showinfo("Analysis", "Report completed")