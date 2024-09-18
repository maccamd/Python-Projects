import os
import pandas as pd
import glob
import fnmatch
import xlsxwriter
import openpyxl as xl
from openpyxl import Workbook
from openpyxl import load_workbook

fgo = r"C:/Users/m0082668/OneDrive - MAHLE\MOTA\Car 1 VA16 FGO\Test Validation FGO.xlsm"
uxb = r"C:/Users/m0082668/OneDrive - MAHLE\MOTA\Car 2 OW17 UXB\Test Validation UXB.xlsx"
llc = r"C:/Users/m0082668/OneDrive - MAHLE\MOTA\Car 3 VA17 LLC\Test Validation LLC.xlsx"
cwf = r"C:/Users/m0082668/OneDrive - MAHLE\MOTA\Car 4 OW67 CWF\Test Validation CWF.xlsx"
zds = r"C:/Users/m0082668/OneDrive - MAHLE\MOTA\Car 5 RV67 ZDS\Test Validation ZDS.xlsx"
kfx = r"C:/Users/m0082668/OneDrive - MAHLE\MOTA\Car 6 MW66 KFX\Test Validation KFX.xlsx"
oyz = r"C:/Users/m0082668/OneDrive - MAHLE\MOTA\Car 7 HJ18 OYZ\Test Validation OYZ.xlsx"
ntk = r"C:/Users/m0082668/OneDrive - MAHLE\MOTA\Car 8 GY18 NTK\Test Validation NTK.xlsx"
oig = r"C:/Users/m0082668/OneDrive - MAHLE\MOTA\Car 9 OIG6607\Test Validation OIG.xlsx"
uaa = r"C:/Users/m0082668/OneDrive - MAHLE\MOTA\Car 10 WK66 UAA\Test Validation UAA.xlsx"
gyc = r"C:/Users/m0082668/OneDrive - MAHLE\MOTA\Car 11 BW66 GYC\Test Validation GYC.xlsx"

def report(filename, car):
    path = r"MOTA Test Output Wk29.xlsx"
    df = pd.read_excel(filename)
    df1 = df.set_index('Test ID', inplace=True)
    df1 = df.drop(columns= ['PM', 'Violation', 'Limit', 'Enviromental', 'INCA', 'RL', 'Emissions Present'], axis=1)
    df1 = df1.rename(columns={"Wk No": "Wk"})
    df2 = df1[(df1.Wk == 29)]
    #print(df2)
    with pd.ExcelWriter(
        path,
        mode="a",
        engine="openpyxl",
        if_sheet_exists="replace",) as writer:
        df2.to_excel(writer, sheet_name=car)

def tidy(file):
    wb = load_workbook(file)
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
        wb.save(file)
    
        
wb = Workbook()
wb.save(filename='MOTA Test Output Wk29.xlsx')
file = r"MOTA Test Output Wk29.xlsx"

#report(fgo,'FGO')
#report(uxb, 'UXB')
#report(llc, 'LLC')
#report(cwf, 'CWF')
#report(zds, 'ZDS')
#report(kfx,'KFX')
report(oyz, 'OYZ')
#report(ntk, 'NTK')
#report(oig, 'OIG')
report(uaa, 'UAA')
report(gyc, 'GYC')
tidy(file)

