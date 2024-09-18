#VDC Extra Analysis Package
"""
Report to drive info out into one report so excel investigation isnt needed and looking into multiple things.
Things TO Report:
REESS Balance over cycle
Filter Flow over Cycle
CVS Volume over cycle

"""
from tkinter.constants import END
import openpyxl as xl
from openpyxl.utils import get_column_letter
import tkinter as tk
from tkinter import Image, filedialog
from tkinter import messagebox
import matplotlib.cbook as cbook
import matplotlib as image
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages
from matplotlib.cbook import get_sample_data
from matplotlib.offsetbox import  OffsetImage
from PIL import Image

#----tkinter----#

root = tk.Tk()
root.withdraw()

#----Functions----#

file_path = filedialog.askopenfilename() #tkinter method
path = open(file_path, "rb") #rb is reading in binary
print(file_path)

wb = xl.load_workbook(path)

sheet = wb['Summary']
filename = sheet['B1'].value
filecut = filename[:-20]

sheet = wb['ContinuousData']  #wb['Continous'] in unprocessed reports     #wb.active is sheet the file opens on

actspeed = "" #Empty string that actual speed will go into
YokoClamp1 = ""
YokoClamp2 = ""


logo = Image.open('MPTLogo.png')

#loops to find columns in the continous data
for row_cells in sheet.iter_rows(min_row=1, max_row=1):
    for cell in row_cells:
        if 'DA Actual Speed' in cell.value or 'SpeedFeedback' in cell.value:
            actspeed = get_column_letter(cell.column)
            
for row_cells in sheet.iter_rows(min_row=1, max_row=1):
        for cell in row_cells:
            if 'Phase' in cell.value:
                phasenum = get_column_letter(cell.column)

for row_cells in sheet.iter_rows(min_row=1, max_row=1):
        for cell in row_cells:
            if 'Yokogawa wt5000REESSCurrent' in cell.value:
                YokoClamp1 = get_column_letter(cell.column)
                
for row_cells in sheet.iter_rows(min_row=1, max_row=1):
        for cell in row_cells:
            if 'Yokogawa wt5000REESSCurrent2' in cell.value:
                YokoClamp2 = get_column_letter(cell.column)

time = [sheet['B' + str(row)].value for row in range(3, sheet.max_row + 1)] #PlotTime Column G for RDEC 2 unprocessed reports
actual_speed = [sheet[actspeed + str(row)].value for row in range(3, sheet.max_row + 1)] #DA Actual Speed
phase = [sheet[phasenum + str(row)].value for row in range(3, sheet.max_row + 1)] #phase
current_clamp1 = [sheet[YokoClamp1 + str(row)].value for row in range(3, sheet.max_row + 1)] #Current Clamp 1
current_clamp2 = [sheet[YokoClamp2 + str(row)].value for row in range(3, sheet.max_row + 1)] #Current Clamp 2

plt.plot(time, actual_speed)
plt.show()