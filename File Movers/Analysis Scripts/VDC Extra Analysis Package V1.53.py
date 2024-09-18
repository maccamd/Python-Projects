#VDC Extra Analysis Package V1.53
##### 
'''
Default column letter of K is usetting finding the coloumn when it exists.
Written By Sam McDonald on 16/03/2023
'''
#####

#Declare Packages
import string
import PySimpleGUI as sg
import openpyxl as xl
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages
from PIL import Image
import time
import pandas as pd
import os

#----Open GUI ----#

layout = [  [sg.Text('Select test file and report destination')],
            [sg.Text('Test Report'), sg.InputText(), sg.FileBrowse()],
            [sg.Text('Save Report'), sg.InputText(), sg.FolderBrowse()],
            [sg.Checkbox('ICE', default=False), sg.Checkbox('EV',default=False)],
            [sg.OK(), sg.Cancel()]]

window = sg.Window('Open Test Report and Save Analysis Report', layout)

event, values = window.read()
window.close()
file_path, folder_path, ice_on, ev_on = values[0], values[1], values[2], values[3]
folder_path_comp = folder_path + '/'

#----Data Gathering----#
path = open(file_path, "rb") #rb is reading in binary
print(file_path)

wb = xl.load_workbook(path)

sheet = wb['Summary']   #inital data gathering, filename
filename = sheet['B1'].value 
filecut = filename[:-20] #takes out last 20 characters so project vehiclename and testid are left in theory

sheet = wb['ContinuousData']  #wb['Continous'] in unprocessed reports     #wb.active is sheet the file opens on

logo = Image.open('MPTLogo.png') # for plotting purposes


#Variable Declare Column title
phase = "Phase" #Phase Number
hioki1 = 'Hioki 3390FundamentalIdc1'
hioki2 = 'Hioki 3390FundamentalIdc2'
hioki3 = 'Hioki 3390FundamentalIdc3'
hioki4 = 'Hioki 3390FundamentalIdc4'
tailpipeCO2Rate = 'TailpipeCO2Rate'
diluteCO2Rate = 'DiluteCO2Rate'
precatCO2Rate = 'PrecatCO2Rate'
tailpipeCOConc = 'TailpipeCOCorrConc'
diluteCOConc = 'DiluteCOCorrConc'
precatCOConc = 'PrecatCOCorrConc'
tailpipeTHCConc = 'TailpipeTHCCorrConc'
diluteTHCConc = 'DiluteTHCCorrConc'
precatTHCConc = 'PrecatTHCCorrConc'
tailpipeNOxConc = 'TailpipeNOXCorrConc'
diluteNOConc = 'DiluteNOCorrConc'
preatNOxConc = 'PrecatNOXCorrConc'
tailpipeCH4Conc = 'TailpipeCH4CorrConc'
diluteCH4Conc = 'DiluteCH4CorrConc'
precatCH4Conc = 'PrecatCH4CorrConc'

#loops to find column id in the continous data
for row_cells in sheet.iter_cols(min_row=1, max_row=1):
    for cell in row_cells:
        
        if 'DA Actual Speed' in cell.value or 'SpeedFeedback' in cell.value:
            actSpeed = cell.column_letter
            
        if cell.value == phase:
            phaseNum = cell.column_letter
#Finding Current Clamps            
        if cell.value == hioki1: 
            dcClamp1 = cell.column_letter
        if cell.value == hioki2: 
            dcClamp2 = cell.column_letter
        if cell.value == hioki3:
            dcClamp3 = cell.column_letter
        if cell.value == hioki4:
            dcClamp4 = cell.column_letter
#Finding CVS & SAO Data                
        if 'FilterPressureDrop' in cell.value:
            filterDrop = cell.column_letter
            
        if 'CVSFlowRate' in cell.value:
            cvsFlow = cell.column_letter
            
        if 'SAODilutionAirFlow' in cell.value:
            saoAirFlow = cell.column_letter

#Finding Gas Compounds, CO2, CO, THC, NOx, CH4             
        if cell.value == tailpipeCO2Rate:
            tailpipeCO2Col = cell.column_letter
        if cell.value == tailpipeCO2Rate:
            diluteCO2Col = cell.column_letter
        if cell.value == precatCO2Rate:
            precatCO2Col = cell.column_letter        
        if cell.value == tailpipeCOConc:
            tailpipeCOcol = cell.column_letter
        if cell.value == diluteCOConc:
            diluteCOcol = cell.column_letter
        if cell.value == precatCOConc:
            precatCOcol = cell.column_letter


        if cell.value == tailpipeTHCConc:
            tailpipeTHCcol = cell.column_letter
        if cell.value == diluteTHCConc:
            diluteTHCcol = cell.column_letter 
        if cell.value == precatTHCConc:
            precatTHCcol = cell.column_letter
        if cell.value == tailpipeNOxConc:
            tailpipeNOxcol = cell.column_letter
        if cell.value == diluteNOConc:
            diluteNOxcol = cell.column_letter
        if cell.value == preatNOxConc:
            precatNOxcol = cell.column_letter
        if cell.value == tailpipeCH4Conc:
            tailpipeCH4col = cell.column_letter
        if cell.value == diluteCH4Conc:
            diluteCH4col = cell.column_letter
        if cell.value == precatCH4Conc:
            precatCH4col = cell.column_letter


#Dyno Infomation
        if 'ChDyno Force Vehicle' in cell.value:
            dynoForce = cell.column_letter
        if'ChDyno Power' in cell.value:
            dynoPower = cell.column_letter
                     
#gets data out of columns above and puts into variables
cycle_time = [sheet['B' + str(row)].value for row in range(3, sheet.max_row + 1)] #PlotTime Column G for RDEC 2 unprocessed reports
actual_speed = [sheet[actSpeed + str(row)].value for row in range(3, sheet.max_row + 1)] #DA Actual Speed
phase = [sheet[phaseNum + str(row)].value for row in range(3, sheet.max_row + 1)] #phase
filterPressure = [sheet[filterDrop + str(row)].value for row in range(3, sheet.max_row + 1)] #Filter Pressure Drop
cvsFlow = [sheet[cvsFlow + str(row)].value for row in range(3, sheet.max_row + 1)] #CVS Flow Rate
exhaustFlow = [sheet[saoAirFlow + str(row)].value for row in range(3, sheet.max_row + 1)] #SAO Air Flow
dynoForceData = [sheet[dynoForce + str(row)].value for row in range(3, sheet.max_row + 1)] #Dyno Force Data
dynoPowerData = [sheet[dynoPower + str(row)].value for row in range(3, sheet.max_row + 1)] #Dyno Power Data

#conditionals to check there is data available or not, empty string if not so plotting gets skipped
if dcClamp1 != []:
    
    current_clamp1 = [sheet[dcClamp1 + str(row)].value for row in range(3, sheet.max_row + 1)] #Current Clamp 1
else:
    current_clamp1 = [2000]

if dcClamp2 != []:
    
    current_clamp2 = [sheet[dcClamp2 + str(row)].value for row in range(3, sheet.max_row + 1)] #Current Clamp 2
else:
    current_clamp3 = [2000]

if dcClamp3 != []:
    
    current_clamp3 = [sheet[dcClamp3 + str(row)].value for row in range(3, sheet.max_row + 1)] #Current Clamp 3
else:
    current_clamp3 = [2000]

if dcClamp4 != []:    
    current_clamp4 = [sheet[dcClamp4 + str(row)].value for row in range(3, sheet.max_row + 1)] #Current Clamp 4
else:
    current_clamp4 = [2000]
    
if tailpipeCO2Col != []:    
    postcatCO2Val = [sheet[tailpipeCO2Col + str(row)].value for row in range(3, sheet.max_row + 1)] #Tailpipe CO2 Rate
else:
    postcatCO2Val = [0]
if diluteCO2Col != []:    
    diluteCO2Val = [sheet[diluteCO2Col + str(row)].value for row in range(3, sheet.max_row + 1)] #Diulte CO2 Rate
else:
    diluteCO2Val = [0]
if precatCO2Col != []:    
    precatCO2Val = [sheet[precatCO2Col + str(row)].value for row in range(3, sheet.max_row + 1)] #Precat CO2 Rate
else:
    precatCO2Val = [0]
    
if tailpipeCOcol != []:    
    postcatCOVal = [sheet[tailpipeCOcol + str(row)].value for row in range(3, sheet.max_row + 1)] #Tailpipe CO Rate
else:
    postcatCOVal = [0]
if diluteCOcol != []:    
    diluteCOVal = [sheet[diluteCOcol + str(row)].value for row in range(3, sheet.max_row + 1)] #Dilute CO Rate
else:
    diluteCOVal = [0]
if precatCOcol != []:    
    precatCOVal = [sheet[precatCOcol + str(row)].value for row in range(3, sheet.max_row + 1)] #Precat CO Rate
else:
    precatCOVal = [0]

if tailpipeTHCcol != []:    
    postcatTHCVal = [sheet[tailpipeTHCcol + str(row)].value for row in range(3, sheet.max_row + 1)] #Tailpipe THC Rate
else:
    postcatTHCVal = [0]
if diluteTHCcol != []:    
    diluteTHCVal = [sheet[diluteTHCcol + str(row)].value for row in range(3, sheet.max_row + 1)] #Dilute THC Rate
else:
    diluteTHCVal = [0]
if precatTHCcol != []:    
    precatTHCVal = [sheet[precatTHCcol + str(row)].value for row in range(3, sheet.max_row + 1)] #Precat THC Rate
else:
    precatTHCVal = [0]

if tailpipeNOxcol != []:    
    postcatNOxVal = [sheet[tailpipeNOxcol + str(row)].value for row in range(3, sheet.max_row + 1)] #Tailpipe NOx Rate
else:
    postcatNOxVal = [0]
if diluteNOxcol != []:    
    diluteNOxVal = [sheet[diluteNOxcol + str(row)].value for row in range(3, sheet.max_row + 1)] #Dilute NOx Rate
else:
    diluteNOxVal = [0]
if precatNOxcol != []:    
    precatNOxVal = [sheet[precatNOxcol + str(row)].value for row in range(3, sheet.max_row + 1)] #Precat NOx Rate
else:
    precatNOxVal = [0]

if tailpipeCH4col != []:    
    postcatCH4Val = [sheet[tailpipeCH4col + str(row)].value for row in range(3, sheet.max_row + 1)] #Tailpipe CH4 Rate
else:
    postcatCH4Val = [0]
if diluteCH4col != []:    
    diluteCH4Val = [sheet[diluteCH4col + str(row)].value for row in range(3, sheet.max_row + 1)] #Dilute CH4 Rate
else:
    diluteCH4Val = [0]
if precatCH4col != []:    
    precatCH4Val = [sheet[precatCH4col + str(row)].value for row in range(3, sheet.max_row + 1)] #Precat CH4 Rate
else:
    precatCH4Val = [0]

#----Plotting Section----#
fig, ax = plt.subplots(figsize=(16.53,11.69)) #Current Vs Trace
plt.subplots_adjust(left=0.04, bottom=0.06, right=0.94, top=0.93, wspace=0.2, hspace=0.23)
newax = fig.add_axes([0.84, 0.881, 0.124, 0.115], anchor='NE',)
newax.imshow(logo)
newax.axis('off')
ax.set_title('REESS Balance over Trace', loc='center')
ax.plot(cycle_time, actual_speed, linewidth=0.9, alpha=0.75)
ax.set_ylabel('Speed (Km/h)')
ax.set_xlabel('Time (s)')
ax.set_xlim([-5,1820])
ax.set_xticks([0, 200, 400, 600, 800, 1000, 1200, 1400, 1600, 1800])
ax.set_ylim([-0.5,140])
ax1 = ax.twinx()
ax1.axhline(y=0, color='g', linestyle='--')
ax1.plot(cycle_time, current_clamp1, color = 'red')
ax1.set_ylabel('Current (A)')
Clamp2OnOff = current_clamp2[0] #Looks for first value
Clamp3OnOff = current_clamp3[0]
Clamp4OnOff = current_clamp4[0]

#section to look at the lines above and if condition isnt met, plotting is skipped
if type(Clamp2OnOff) == float: 
    ax1.plot(cycle_time, current_clamp2, color = 'grey')  
if type(Clamp3OnOff) == float:
    ax1.plot(cycle_time, current_clamp3, color = 'orange')  
if type(Clamp4OnOff) == float:
    ax1.plot(cycle_time, current_clamp4, color = 'purple')
    
fig2, ax2 = plt.subplots(figsize=(16.53,11.69)) #Filter Pressure Drop Vs Trace
plt.subplots_adjust(left=0.04, bottom=0.06, right=0.94, top=0.93, wspace=0.2, hspace=0.23)
newax = fig2.add_axes([0.84, 0.881, 0.124, 0.115], anchor='NE',)
newax.imshow(logo)
newax.axis('off')
ax2.set_title('Filter Pressure Drop over Trace', loc='center')
ax2.plot(cycle_time, actual_speed, linewidth=0.9, alpha=0.75)
ax2.set_ylabel('Speed (Km/h)')
ax2.set_xlabel('Time (s)')
ax2.set_xlim([-5,1820])
ax2.set_xticks([0, 200, 400, 600, 800, 1000, 1200, 1400, 1600, 1800])
ax2.set_ylim([-0.5,140])
ax3 = ax2.twinx()
ax3.plot(cycle_time, filterPressure, color = 'red')
ax3.set_ylabel('Filter Pressure Drop (Pa)')

fig3, ax8 = plt.subplots(figsize=(16.53,11.69)) #CVS Flow Rate Vs Trace
plt.subplots_adjust(left=0.04, bottom=0.06, right=0.94, top=0.93, wspace=0.2, hspace=0.23)
newax = fig3.add_axes([0.84, 0.881, 0.124, 0.115], anchor='NE',)
newax.imshow(logo)
newax.axis('off')
ax8.set_title('CVS Flow Rate over Trace', loc='center')
ax8.plot(cycle_time, actual_speed, linewidth=0.9, alpha=0.75)
ax8.set_ylabel('Speed (Km/h)')
ax8.set_xlabel('Time (s)')
ax8.set_xlim([-5,1820])
ax8.set_xticks([0, 200, 400, 600, 800, 1000, 1200, 1400, 1600, 1800])
ax8.set_ylim([-0.5,140])
ax5 = ax8.twinx()
ax5.plot(cycle_time, cvsFlow, color = 'red',linewidth=0.5)
ax5.set_ylabel('CVS Flow Rate (m3/min)')

fig4, ax6 = plt.subplots(figsize=(16.53,11.69)) #SAO Flow Rate Vs Trace
plt.subplots_adjust(left=0.04, bottom=0.06, right=0.94, top=0.93, wspace=0.2, hspace=0.23)
newax = fig4.add_axes([0.84, 0.881, 0.124, 0.115], anchor='NE',)
newax.imshow(logo)
newax.axis('off')
ax6.set_title('SAO Flow Rate over Trace', loc='center')
ax6.plot(cycle_time, actual_speed, linewidth=0.9, alpha=0.75)
ax6.set_ylabel('Speed (Km/h)')
ax6.set_xlabel('Time (s)')
ax6.set_xlim([-5,1820])
ax6.set_xticks([0, 200, 400, 600, 800, 1000, 1200, 1400, 1600, 1800])
ax6.set_ylim([-0.5,140])
ax7 = ax6.twinx()
ax7.plot(cycle_time, exhaustFlow, color = 'red',linewidth=1)
ax7.set_ylabel('SAO Flow Rate (m3/s)')

fig5, ax8 = plt.subplots(figsize=(16.53,11.69)) #CO2 Vs Trace
plt.subplots_adjust(left=0.04, bottom=0.06, right=0.94, top=0.93, wspace=0.2, hspace=0.23)
newax = fig5.add_axes([0.84, 0.881, 0.124, 0.115], anchor='NE',)
newax.imshow(logo)
newax.axis('off')
ax8.set_title('CO2 Rate over trace', loc='center')
ax8.plot(cycle_time, actual_speed, linewidth=0.9, alpha=0.75)
ax8.set_ylabel('Speed (Km/h)')
ax8.set_xlabel('Time (s)')
ax8.set_xlim([-5,1820])
ax8.set_xticks([0, 200, 400, 600, 800, 1000, 1200, 1400, 1600, 1800])
ax8.set_ylim([-0.5,140])
ax9 = ax8.twinx()
ax9.set_ylabel('CO2 Rate (g/s)')
fig5.legend(ncol=3, loc='upper center')
if tailpipeCO2Col != 'K':
    ax9.plot(cycle_time, postcatCO2Val, color = 'red',linewidth=1, label='Postcat Line')
if diluteCO2Col !='K':
     ax9.plot(cycle_time, diluteCO2Val, color = 'green',linewidth=1, label='Dilute Line')
if precatCO2Col != 'K':
    ax9.plot(cycle_time, precatCO2Val, color = 'orange',linewidth=1, label='Precat Line')      



fig6, ax10 = plt.subplots(figsize=(16.53,11.69)) #CO Vs Trace
plt.subplots_adjust(left=0.04, bottom=0.06, right=0.94, top=0.93, wspace=0.2, hspace=0.23)
newax = fig6.add_axes([0.84, 0.881, 0.124, 0.115], anchor='NE',)
newax.imshow(logo)
newax.axis('off')
ax10.set_title('CO Rate over trace', loc='center')
ax10.plot(cycle_time, actual_speed, linewidth=0.9, alpha=0.75)
ax10.set_ylabel('Speed (Km/h)')
ax10.set_xlabel('Time (s)')
ax10.set_xlim([-5,1820])
ax10.set_xticks([0, 200, 400, 600, 800, 1000, 1200, 1400, 1600, 1800])
ax10.set_ylim([-0.5,140])
diluteOnOff = diluteCOVal[0]
tailpipeOnOff = postcatCOVal[0]
precatOnOff = precatCOVal[0]
if type(diluteOnOff or tailpipeOnOff or precatOnOff ) == float:
    ax11 = ax10.twinx()
    ax11.plot(cycle_time, postcatCOVal, color = 'red',linewidth=1, label='Postcat Line')
    ax11.plot(cycle_time, diluteCOVal, color = 'green',linewidth=1, label='Dilute Line')
    ax11.plot(cycle_time, precatCOVal, color = 'orange',linewidth=1, label='Precat Line')
    ax11.set_ylabel('Postcat CO Rate (g/s)')
    fig6.legend(ncol=3, loc='upper center')

fig7, ax12 = plt.subplots(figsize=(16.53,11.69)) #THC Vs Trace
plt.subplots_adjust(left=0.04, bottom=0.06, right=0.94, top=0.93, wspace=0.2, hspace=0.23)
newax = fig7.add_axes([0.84, 0.881, 0.124, 0.115], anchor='NE',)
newax.imshow(logo)
newax.axis('off')
ax12.set_title('THC Rate over trace', loc='center')
ax12.plot(cycle_time, actual_speed, linewidth=0.9, alpha=0.75)
ax12.set_ylabel('Speed (Km/h)')
ax12.set_xlabel('Time (s)')
ax12.set_xlim([-5,1820])
ax12.set_xticks([0, 200, 400, 600, 800, 1000, 1200, 1400, 1600, 1800])
ax12.set_ylim([-0.5,140])
diluteOnOff = diluteTHCVal[0]
tailpipeOnOff = postcatTHCVal[0]
precatOnOff = precatTHCVal[0]
if type(diluteOnOff or tailpipeOnOff or precatOnOff ) == float:
    ax13 = ax12.twinx()
    ax13.plot(cycle_time, postcatTHCVal, color = 'red',linewidth=1, label='Postcat Line')
    ax13.plot(cycle_time, diluteTHCVal, color = 'green',linewidth=1, label='Dilute Line')
    ax13.plot(cycle_time, precatTHCVal, color = 'orange',linewidth=1, label='Precat Line')
    ax13.set_ylabel('Postcat THC Rate (g/s)')
    fig7.legend(ncol=3, loc='upper center')

fig8, ax14 = plt.subplots(figsize=(16.53,11.69)) #NOx Vs Trace
plt.subplots_adjust(left=0.04, bottom=0.06, right=0.94, top=0.93, wspace=0.2, hspace=0.23)
newax = fig8.add_axes([0.84, 0.881, 0.124, 0.115], anchor='NE',)
newax.imshow(logo)
newax.axis('off')
ax14.set_title('NOx Rate over trace', loc='center')
ax14.plot(cycle_time, actual_speed, linewidth=0.9, alpha=0.75)
ax14.set_ylabel('Speed (Km/h)')
ax14.set_xlabel('Time (s)')
ax14.set_xlim([-5,1820])
ax14.set_xticks([0, 200, 400, 600, 800, 1000, 1200, 1400, 1600, 1800])
ax14.set_ylim([-0.5,140])
diluteOnOff = diluteNOxVal[0]
tailpipeOnOff = postcatNOxVal[0]
precatOnOff = precatNOxVal[0]
if type(diluteOnOff or tailpipeOnOff or precatOnOff ) == float:
    ax15 = ax14.twinx()
    ax15.plot(cycle_time, postcatNOxVal, color = 'red',linewidth=1, label='Postcat Line')
    ax15.plot(cycle_time, diluteNOxVal, color = 'green',linewidth=1, label='Dilute Line')
    ax15.plot(cycle_time, precatNOxVal, color = 'orange',linewidth=1, label='Precat Line')
    ax15.set_ylabel('Postcat NOx Rate (g/s)')
    fig8.legend(ncol=3, loc='upper center')

fig9, ax16 = plt.subplots(figsize=(16.53,11.69)) #CH4 Vs Trace
plt.subplots_adjust(left=0.04, bottom=0.06, right=0.94, top=0.93, wspace=0.2, hspace=0.23)
newax = fig9.add_axes([0.84, 0.881, 0.124, 0.115], anchor='NE',)
newax.imshow(logo)
newax.axis('off')
ax16.set_title('CH4 Rate over trace', loc='center')
ax16.plot(cycle_time, actual_speed, linewidth=0.9, alpha=0.75)
ax16.set_ylabel('Speed (Km/h)')
ax16.set_xlabel('Time (s)')
ax16.set_xlim([-5,1820])
ax16.set_xticks([0, 200, 400, 600, 800, 1000, 1200, 1400, 1600, 1800])
ax16.set_ylim([-0.5,140])
diluteOnOff = diluteCH4Val[0]
tailpipeOnOff = postcatCH4Val[0]
precatOnOff = precatCH4Val[0]
if type(diluteOnOff or tailpipeOnOff or precatOnOff ) == float:
    ax17 = ax16.twinx()
    ax17.plot(cycle_time, postcatCH4Val, color = 'red',linewidth=1, label='Postcat Line')
    ax17.plot(cycle_time, diluteCH4Val, color = 'green',linewidth=1, label='Dilute Line')
    ax17.plot(cycle_time, precatCH4Val, color = 'orange',linewidth=1, label='Precat Line')
    ax17.set_ylabel('Postcat CH4 Rate (g/s)')
    fig9.legend(ncol=3, loc='upper center')

fig10, ax18 = plt.subplots(figsize=(16.53,11.69)) #Dyno Force Vs Trace
plt.subplots_adjust(left=0.04, bottom=0.06, right=0.94, top=0.93, wspace=0.2, hspace=0.23)
newax = fig10.add_axes([0.84, 0.881, 0.124, 0.115], anchor='NE',)
newax.imshow(logo)
newax.axis('off')
ax18.set_title('Dyno Force over trace', loc='center')
ax18.plot(cycle_time, actual_speed, linewidth=0.9, alpha=0.75)
ax18.set_ylabel('Speed (Km/h)')
ax18.set_xlabel('Time (s)')
ax18.set_xlim([-5,1820])
ax18.set_xticks([0, 200, 400, 600, 800, 1000, 1200, 1400, 1600, 1800])
ax18.set_ylim([-0.5,140])
ax19 = ax18.twinx()
ax19.plot(cycle_time, dynoForceData, color = 'red',linewidth=1)
ax19.set_ylabel('Dyno Force (N)')

fig11, ax20 = plt.subplots(figsize=(16.53,11.69)) #Dyno Power Vs Trace
plt.subplots_adjust(left=0.04, bottom=0.06, right=0.94, top=0.93, wspace=0.2, hspace=0.23)
newax = fig11.add_axes([0.84, 0.881, 0.124, 0.115], anchor='NE',)
newax.imshow(logo)
newax.axis('off')
ax20.set_title('Dyno Power over trace', loc='center')
ax20.plot(cycle_time, actual_speed, linewidth=0.9, alpha=0.75)
ax20.set_ylabel('Speed (Km/h)')
ax20.set_xlabel('Time (s)')
ax20.set_xlim([-5,1820])
ax20.set_xticks([0, 200, 400, 600, 800, 1000, 1200, 1400, 1600, 1800])
ax20.set_ylim([-0.5,140])
ax21 = ax20.twinx()
ax21.plot(cycle_time, dynoPowerData, color = 'red',linewidth=1)
ax21.set_ylabel('Dyno Power (mW)')

def save_image(filename):
    
    # PdfPages is a wrapper around pdf file so there is no clash and create files with no error.
    p = PdfPages(filename)
    # get_fignums Return list of existing figure numbers
    fig_nums = plt.get_fignums()  
    figs = [plt.figure(n) for n in fig_nums]
    # iterating over the numbers in list
    for fig in figs: 
        # and saving the files
        fig.savefig(p, format='pdf') 
    # close the object
    p.close() 

filename = folder_path_comp + filecut + " Analysis.pdf"
save_image(filename)
print('Analysis Complete')