import PySimpleGUI as sg
import openpyxl as xl
import os
from openpyxl.utils import get_column_letter
import matplotlib.cbook as cbook
import matplotlib as image
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages
from matplotlib.cbook import get_sample_data
from matplotlib.offsetbox import  OffsetImage
from PIL import Image

#Driver Trace Analysis Tool
#Version 2 of code

#----PySimpleGUI----#

layout = [  [sg.Text('Select test file and report destination')],
            [sg.Text('Test Report'), sg.InputText(), sg.FileBrowse()],
            [sg.Text('Save Report'), sg.InputText(), sg.FolderBrowse()],
            [sg.OK(), sg.Cancel()]]

window = sg.Window('Open Test Report and Save Analysis Report', layout)

event, values = window.read()
window.close()
file_path, folder_path = values[0], values[1]
folder_path_comp = folder_path + '/'
my_path = os.path.abspath(folder_path_comp)
print(my_path)

class Plotting:
    
    def plot():
        with PdfPages(filecut + ".pdf") as pdf:
            global filename, driver, errorcnt1, errorcnt2, errorcnt3, errorcnt4, logo, phase1_time, phase1_target_speed, phase1_actual_speed, phase1_upper_boundary, phase1_lower_boundary, phase2_target_speed, phase2_actual_speed, phase2_upper_boundary, phase2_lower_boundary
            fig, axs1 = plt.subplots(2, figsize=(11.69,8.27))
            plt.subplots_adjust(left=0.05, bottom=0.06, right=0.97, top=0.93, wspace=0.2, hspace=0.23)
            fig.suptitle(filename, x=0.1, y=0.95, horizontalalignment='left', verticalalignment='top')
            newax = fig.add_axes([0.84, 0.881, 0.124, 0.115], anchor='NE',)
            newax.imshow(logo)
            newax.axis('off')    
            axs1[0].set_title('Phase 1')
            axs1[0].set_xlabel('Time (s)')
            axs1[0].set_ylabel('Speed (Km/h)')
            axs1[0].text(0, 66, 'Driver: ' + driver)
            axs1[0].text(0, 63, 'Error Count: ' + errorcnt1 + ',  Error Time: ' + errortime1)
            axs1[0].text(0, 60.5, 'Violation Count: ' + violationcnt1 + ',  Violation Time: ' + violationtime1 + ', IWR Phase 1: ' + iwr1 + '%')
            axs1[0].text(410, 63, 'IWR Test: ' + iwrtest + '%')
            axs1[0].plot(phase1_time,phase1_target_speed, label="Target Speed", linewidth= 0.75) # line width unreadable below 0.5
            axs1[0].plot(phase1_time,phase1_actual_speed, label="Actual Speed",linewidth= 0.75)
            axs1[0].plot(phase1_time,phase1_upper_boundary, linestyle= '--',linewidth= 0.75, color= 'grey')
            axs1[0].plot(phase1_time,phase1_lower_boundary, linestyle= '--',linewidth= 0.75, color= 'grey')
            axs1[0].legend()
            axs1[0].grid(linestyle = '--', linewidth= 0.5)
            axs1[0].margins(0)
            axs1[0].set_ylim([-0.3,60])

            axs1[1].set_title('Phase 2')
            axs1[1].set_xlabel('Time (s)')
            axs1[1].set_ylabel('Speed (Km/h)')
            axs1[1].text(589, 84, 'Error Count: ' + errorcnt2 + ',  Error Time: ' + errortime2)
            axs1[1].text(589, 80.5, 'Violation Count: ' + violationcnt2 + ',  Violation Time: ' + violationtime2 + ', IWR Phase 2: ' + iwr2 + '%')
            axs1[1].plot(phase2_time,phase2_target_speed, label="Target Speed", linewidth= 0.75,)
            axs1[1].plot(phase2_time,phase2_actual_speed, label="Actual Speed", linewidth= 0.75)
            axs1[1].plot(phase2_time,phase2_upper_boundary, linestyle= '--',linewidth= 0.75, color= 'grey')
            axs1[1].plot(phase2_time,phase2_lower_boundary, linestyle= '--',linewidth= 0.75, color= 'grey')
            axs1[1].legend()
            axs1[1].grid(linestyle = '--', linewidth= 0.5)
            axs1[1].margins(0)
            axs1[1].set_ylim([-0.3,80])
            pdf.savefig()

            #Figure 2
            fig2, axs2 = plt.subplots(2, figsize=(11.96,8.27))
            plt.subplots_adjust(left=0.05, bottom=0.06, right=0.97, top=0.93, wspace=0.2, hspace=0.23)
            fig.suptitle(filename)
            newax = fig2.add_axes([0.84, 0.881, 0.124, 0.115], anchor='NE',)
            newax.imshow(logo)
            newax.axis('off')    
            axs2[0].set_title('Phase 3')
            axs2[0].set_xlabel('Time (s)')
            axs2[0].set_ylabel('Speed (Km/h)')
            axs2[0].text(1022, 106, 'Error Count: ' + errorcnt3 + ',  Error Time: ' + errortime3)
            axs2[0].text(1022, 101.5, 'Violation Count: ' + violationcnt3 + ',  Violation Time: ' + violationtime3 + ', IWR Phase 3: ' + iwr3 + '%')
            axs2[0].plot(phase3_time,phase3_target_speed, label="Target Speed", linewidth= 0.75)
            axs2[0].plot(phase3_time,phase3_actual_speed, label="Actual Speed", linewidth= 0.75)
            axs2[0].plot(phase3_time,phase3_upper_boundary, linestyle= '--',linewidth= 0.75, color= 'grey')
            axs2[0].plot(phase3_time,phase3_lower_boundary, linestyle= '--',linewidth= 0.75, color= 'grey')
            axs2[0].legend()
            axs2[0].grid(linestyle = '--', linewidth= 0.5)
            axs2[0].margins(0)
            axs2[0].set_xlim([1022,1477])
            axs2[0].set_ylim([-0.3,101])

            axs2[1].set_title('Phase 4')
            axs2[1].set_xlabel('Time (s)')
            axs2[1].set_ylabel('Speed (Km/h)')
            axs2[1].text(1477, 142, 'Error Count: ' + errorcnt4 + ',  Error Time: ' + errortime4)
            axs2[1].text(1477, 136.8, 'Violation Count: ' + violationcnt4 + ',  Violation Time: ' + violationtime4 + ', IWR Phase 4: ' + iwr4 + '%')
            axs2[1].plot(phase4_time,phase4_target_speed, label="Target Speed", linewidth= 0.75)
            axs2[1].plot(phase4_time,phase4_actual_speed, label="Actual Speed", linewidth= 0.75)
            axs2[1].plot(phase4_time,phase4_upper_boundary, linestyle= '--',linewidth= 0.75, color= 'grey')
            axs2[1].plot(phase4_time,phase4_lower_boundary, linestyle= '--',linewidth= 0.75, color= 'grey')
            axs2[1].legend()
            axs2[1].grid(linestyle = '--', linewidth= 0.5)
            axs2[1].margins(0)
            axs2[1].set_xlim([1477,1820])
            axs2[1].set_ylim([-0.3,135])
            pdf.savefig(os.path.join(my_path, file_path))
            pdf.close()

        #plt.rcParams["savefig.directory"] = my_path
        filename = PdfPages(filecut + ".pdf")
        filename.savefig(fig, fig2)
        filename.savefig(fig2)
        filename.close()

path = open(file_path, "rb") #rb is reading in binary
print(file_path)

#----Opens Excel sheet----#
wb = xl.load_workbook(path)
sheet = wb['ContinuousData']  #wb['ContinousData'] in unprocessed reports     #wb.active is sheet the file opens on

logo = Image.open('MPTLogo.png')

#----Loop to get data out of sheet----#

for row_cells in sheet.iter_rows(min_row=1, max_row=1):
    for cell in row_cells:
        if 'Phase' in cell.value:
            phasenum = cell.column_letter
        if 'ScheduledSpeed' in cell.value or 'DA Schedule Speed' in cell.value:
            tarspeed = cell.column_letter
        if 'SpeedFeedback' in cell.value or 'DA Actual Speed' in cell.value:
            actspeed = cell.column_letter
        if 'UpperTolerance' in cell.value:
            uppertol = cell.column_letter
        if 'LowerTolerance' in cell.value:
            lowertol = cell.column_letter
        

time = [sheet['B' + str(row)].value for row in range(3, sheet.max_row + 1)] #PlotTime Column G for RDEC 2 unprocessed reports
target_speed = [sheet[tarspeed + str(row)].value for row in range(3, sheet.max_row + 1)] #DA Schedule Speed
actual_speed = [sheet[actspeed + str(row)].value for row in range(3, sheet.max_row + 1)] #DA Actual Speed
upper_boundary = [sheet[uppertol + str(row)].value for row in range(3, sheet.max_row + 1)] #UpperTolerance
lower_boundary = [sheet[lowertol + str(row)].value for row in range(3, sheet.max_row + 1)] #LowerTolerance
phase = [sheet[phasenum + str(row)].value for row in range(3, sheet.max_row + 1)] #phase

phase1_time = time[0:5890]
phase1_target_speed = target_speed[0:5890]
phase1_actual_speed = actual_speed[0:5890]
phase1_upper_boundary = upper_boundary[0:5890]
phase1_lower_boundary = lower_boundary[0:5890]

phase2_time = time[5891:10220]
phase2_target_speed = target_speed[5891:10220]
phase2_actual_speed = actual_speed[5891:10220] 
phase2_upper_boundary = upper_boundary[5891:10220]
phase2_lower_boundary = lower_boundary[5891:10220]

phase3_time = time[10221:14770]
phase3_target_speed = target_speed[10221:14770]
phase3_actual_speed = actual_speed[10221:14770]
phase3_upper_boundary = upper_boundary[10221:14770]
phase3_lower_boundary = lower_boundary[10221:14770]

phase4_time = time[14771:]
phase4_target_speed = target_speed[14771:]
phase4_actual_speed = actual_speed[14771:]
phase4_upper_boundary = upper_boundary[14771:]
phase4_lower_boundary = lower_boundary[14771:]

sheet = wb['Summary']
filename = sheet['B1'].value
filecut = filename[:-20]
driver = sheet['AD3'].value
cycle = sheet['D2'].value
iwrtest = str(round(sheet['T76'].value, 2))

sheet = wb['Phase1']
errorcnt1 = str(sheet['D16'].value)
errortime1 = str(round(sheet['H16'].value, 3))
violationcnt1 = str(sheet['L16'].value)
iwr1 = str(round(sheet['T14'].value, 2))
violationtime1 = str(round(sheet['T16'].value, 3))

sheet = wb['Phase2']
errorcnt2 = str(sheet['D16'].value)
errortime2 = str(round(sheet['H16'].value, 3))
violationcnt2 = str(sheet['L16'].value)
iwr2 = str(round(sheet['T14'].value, 2))
violationtime2 = str(round(sheet['T16'].value, 3))

sheet = wb['Phase3']
errorcnt3 = str(sheet['D16'].value)
errortime3 = str(round(sheet['H16'].value, 3))
violationcnt3 = str(sheet['L16'].value)
iwr3 = str(round(sheet['T14'].value, 2))
violationtime3 = str(round(sheet['T16'].value, 3))

sheet = wb['Phase4']
errorcnt4 = str(sheet['D16'].value)
errortime4 = str(round(sheet['H16'].value, 3))
violationcnt4 = str(sheet['L16'].value)
iwr4 = str(round(sheet['T14'].value, 2))
violationtime4 = str(round(sheet['T16'].value, 3))

test = Plotting.plot()