#VDC Extra Analysis Package V1.5
##### 
'''
Should be final copy before optimisation rework
Written By Sam McDonald on 16/03/2023
'''
#####

#Declare Packages
import string
import PySimpleGUI as sg
import openpyxl as xl
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages
from PIL import Image
import time
import pandas as pd
import os

#----Open GUI ----#

layout = [  [sg.Text('Select test file and report destination')],
            [sg.Text('Test Report'), sg.InputText(), sg.FileBrowse()],
            [sg.Text('Save Report'), sg.InputText(), sg.FolderBrowse()],
            [sg.Checkbox('ICE', default=False), sg.Checkbox('EV',default=False)],
            [sg.OK(), sg.Cancel()]]

window = sg.Window('Open Test Report and Save Analysis Report', layout)

event, values = window.read()
window.close()
file_path, folder_path, ice_on, ev_on = values[0], values[1], values[2], values[3]
folder_path_comp = folder_path + '/'

#----Data Gathering----#
path = open(file_path, "rb") #rb is reading in binary
print(file_path)

wb = xl.load_workbook(path)

sheet = wb['Summary']   #inital data gathering, filename
filename = sheet['B1'].value 
filecut = filename[:-20] #takes out last 20 characters so project vehiclename and testid are left in theory

sheet = wb['ContinuousData']  #wb['Continous'] in unprocessed reports     #wb.active is sheet the file opens on

logo = Image.open('MPTLogo.png') # for plotting purposes


#Variable Declare Column title
phase = "Phase" #Phase Number

#loops to find column id in the continous data
for row_cells in sheet.iter_cols(min_row=1, max_row=1):
    for cell in row_cells:
        
        if 'DA Actual Speed' in cell.value or 'SpeedFeedback' in cell.value:
            actSpeed = cell.column_letter
            
        if cell.value == phase:
            phaseNum = cell.column_letter
#Finding Current Clamps            
        if 'Hioki 3390FundamentalIdc1' in cell.value: 
            dcClamp1 = cell.column_letter
        else:
            dcClamp1 = "K"
                
        if 'Hioki 3390FundamentalIdc2' in cell.value: 
            dcClamp2 = cell.column_letter
        else:
            dcClamp2 = "K"
                
        if 'Hioki 3390FundamentalIdc3' in cell.value:
            dcClamp3 = cell.column_letter
        else:
            dcClamp3 = "K"
                
        if 'Hioki 3390FundamentalIdc4' in cell.value:
            dcClamp4 = cell.column_letter
        else:
            dcClamp4 = "K"
#Finding CVS & SAO Data                
        if 'FilterPressureDrop' in cell.value:
            filterDrop = cell.column_letter
        else:
            filterDrop = "K"
            
        if 'CVSFlowRate' in cell.value:
            cvsFlow = cell.column_letter
            
        if 'SAODilutionAirFlow' in cell.value:
            saoAirFlow = cell.column_letter

#Finding Gas Compounds, CO2, CO, THC, NOx, CH4             
        if 'TailpipeCO2Rate' in cell.value:
            tailpipeCO2flow = cell.column_letter
        if 'DiluteCO2Rate' in cell.value:
            diluteCO2Rate = cell.column_letter
        if 'PrecatCO2Rate' in cell.value:
            precatCO2Rate = cell.column_letter
                    
        if 'TailpipeCOCorrConc' in cell.value:
            tailpipeCOflow = cell.column_letter
        if 'DiluteCOCorrConc' in cell.value:
            diluteCOflow = cell.column_letter
        if 'PrecatCOCorrConc' in cell.value:
            precatCOflow = cell.column_letter

        if 'TailpipeTHCCorrConc' in cell.value:
            tailpipeTHCflow = cell.column_letter
        if 'DiluteTHCCorrConc' in cell.value:
            diluteTHCflow = cell.column_letter
        if 'PrecatTHCCorrConc' in cell.value:
            precatTHCflow = cell.column_letter

        if 'TailpipeNOXCorrConc' in cell.value:
            tailpipeNOxflow = cell.column_letter
        if 'DiluteNOCorrConc' in cell.value:
            diluteNOxflow = cell.column_letter
        if 'PrecatNOXCorrConc' in cell.value:
            precatNOxflow = cell.column_letter

        if 'TailpipeCH4CorrConc' in cell.value:
            tailpipeCH4flow = cell.column_letter
        if 'DiluteCH4CorrConc' in cell.value:
            diluteCH4flow = cell.column_letter
        if 'PrecatCH4CorrConc' in cell.value:
            precatCH4flow = cell.column_letter

#Dyno Infomation
        if 'ChDyno Force Vehicle' in cell.value:
            dynoForce = cell.column_letter
        if'ChDyno Power' in cell.value:
            dynoPower = cell.column_letter
                     
#gets data out of columns above and puts into variables
cycle_time = [sheet['B' + str(row)].value for row in range(3, sheet.max_row + 1)] #PlotTime Column G for RDEC 2 unprocessed reports
actual_speed = [sheet[actSpeed + str(row)].value for row in range(3, sheet.max_row + 1)] #DA Actual Speed
phase = [sheet[phaseNum + str(row)].value for row in range(3, sheet.max_row + 1)] #phase
filterPressure = [sheet[filterDrop + str(row)].value for row in range(3, sheet.max_row + 1)] #Filter Pressure Drop
cvsFlow = [sheet[cvsFlow + str(row)].value for row in range(3, sheet.max_row + 1)] #CVS Flow Rate
exhaustFlow = [sheet[saoAirFlow + str(row)].value for row in range(3, sheet.max_row + 1)] #SAO Air Flow
dynoForceData = [sheet[dynoForce + str(row)].value for row in range(3, sheet.max_row + 1)] #Dyno Force Data
dynoPowerData = [sheet[dynoPower + str(row)].value for row in range(3, sheet.max_row + 1)] #Dyno Power Data

#conditionals to check there is data available or not, empty string if not so plotting gets skipped
if dcClamp1 != []:
    
    current_clamp1 = [sheet[dcClamp1 + str(row)].value for row in range(3, sheet.max_row + 1)] #Current Clamp 1
else:
    current_clamp1 = [2000]

if dcClamp2 != []:
    
    current_clamp2 = [sheet[dcClamp2 + str(row)].value for row in range(3, sheet.max_row + 1)] #Current Clamp 2
else:
    current_clamp3 = [2000]

if dcClamp3 != []:
    
    current_clamp3 = [sheet[dcClamp3 + str(row)].value for row in range(3, sheet.max_row + 1)] #Current Clamp 3
else:
    current_clamp3 = [2000]

if dcClamp4 != []:    
    current_clamp4 = [sheet[dcClamp4 + str(row)].value for row in range(3, sheet.max_row + 1)] #Current Clamp 4
else:
    current_clamp4 = [2000]
    
if tailpipeCO2flow != []:    
    postcatCO2Rate = [sheet[tailpipeCO2flow + str(row)].value for row in range(3, sheet.max_row + 1)] #Tailpipe CO2 Rate
else:
    postcatCO2Rate = [0]
if diluteCO2Rate != []:    
    diluteCO2Val = [sheet[diluteCO2Rate + str(row)].value for row in range(3, sheet.max_row + 1)] #Diulte CO2 Rate
else:
    diluteCO2Val = [0]
if precatCO2Rate != []:    
    precatCO2Val = [sheet[precatCO2Rate + str(row)].value for row in range(3, sheet.max_row + 1)] #Precat CO2 Rate
else:
    precatCO2Val = [0]
    
if tailpipeCOflow != []:    
    postcatCORate = [sheet[tailpipeCOflow + str(row)].value for row in range(3, sheet.max_row + 1)] #Tailpipe CO Rate
else:
    postcatCORate = [0]
if diluteCOflow != []:    
    diluteCORate = [sheet[diluteCOflow + str(row)].value for row in range(3, sheet.max_row + 1)] #Dilute CO Rate
else:
    diluteCORate = [0]
if precatCOflow != []:    
    precatCORate = [sheet[precatCOflow + str(row)].value for row in range(3, sheet.max_row + 1)] #Precat CO Rate
else:
    precatCORate = [0]

if tailpipeTHCflow != []:    
    postcatTHCRate = [sheet[tailpipeTHCflow + str(row)].value for row in range(3, sheet.max_row + 1)] #Tailpipe THC Rate
else:
    postcatTHCRate = [0]
if diluteTHCflow != []:    
    diluteTHCRate = [sheet[diluteTHCflow + str(row)].value for row in range(3, sheet.max_row + 1)] #Dilute THC Rate
else:
    diluteTHCRate = [0]
if precatTHCflow != []:    
    precatTHCRate = [sheet[precatTHCflow + str(row)].value for row in range(3, sheet.max_row + 1)] #Precat THC Rate
else:
    precatTHCRate = [0]

if tailpipeNOxflow != []:    
    postcatNOxRate = [sheet[tailpipeNOxflow + str(row)].value for row in range(3, sheet.max_row + 1)] #Tailpipe NOx Rate
else:
    postcatNOxRate = [0]
if diluteNOxflow != []:    
    diluteNOxRate = [sheet[diluteNOxflow + str(row)].value for row in range(3, sheet.max_row + 1)] #Dilute NOx Rate
else:
    diluteNOxRate = [0]
if precatNOxflow != []:    
    precatNOxRate = [sheet[precatNOxflow + str(row)].value for row in range(3, sheet.max_row + 1)] #Precat NOx Rate
else:
    precatNOxRate = [0]

if tailpipeCH4flow != []:    
    postcatCH4Rate = [sheet[tailpipeCH4flow + str(row)].value for row in range(3, sheet.max_row + 1)] #Tailpipe CH4 Rate
else:
    postcatCH4Rate = [0]
if diluteCH4flow != []:    
    diluteCH4Rate = [sheet[diluteCH4flow + str(row)].value for row in range(3, sheet.max_row + 1)] #Dilute CH4 Rate
else:
    diluteCH4Rate = [0]
if precatCH4flow != []:    
    precatCH4Rate = [sheet[precatCH4flow + str(row)].value for row in range(3, sheet.max_row + 1)] #Precat CH4 Rate
else:
    precatCH4Rate = [0]

#----Plotting Section----#
fig, ax = plt.subplots(figsize=(16.53,11.69)) #Current Vs Trace
plt.subplots_adjust(left=0.04, bottom=0.06, right=0.94, top=0.93, wspace=0.2, hspace=0.23)
newax = fig.add_axes([0.84, 0.881, 0.124, 0.115], anchor='NE',)
newax.imshow(logo)
newax.axis('off')
ax.set_title('REESS Balance over Trace', loc='center')
ax.plot(cycle_time, actual_speed, linewidth=0.9, alpha=0.75)
ax.set_ylabel('Speed (Km/h)')
ax.set_xlabel('Time (s)')
ax.set_xlim([-5,1820])
ax.set_xticks([0, 200, 400, 600, 800, 1000, 1200, 1400, 1600, 1800])
ax.set_ylim([-0.5,140])
ax1 = ax.twinx()
ax1.axhline(y=0, color='g', linestyle='--')
ax1.plot(cycle_time, current_clamp1, color = 'red')
ax1.set_ylabel('Current (A)')
Clamp2OnOff = current_clamp2[0] #Looks for first value
Clamp3OnOff = current_clamp3[0]
Clamp4OnOff = current_clamp4[0]

#section to look at the lines above and if condition isnt met, plotting is skipped
if type(Clamp2OnOff) == float: 
    ax1.plot(cycle_time, current_clamp2, color = 'grey')  
if type(Clamp3OnOff) == float:
    ax1.plot(cycle_time, current_clamp3, color = 'orange')  
if type(Clamp4OnOff) == float:
    ax1.plot(cycle_time, current_clamp4, color = 'purple')
    
fig2, ax2 = plt.subplots(figsize=(16.53,11.69)) #Filter Pressure Drop Vs Trace
plt.subplots_adjust(left=0.04, bottom=0.06, right=0.94, top=0.93, wspace=0.2, hspace=0.23)
newax = fig2.add_axes([0.84, 0.881, 0.124, 0.115], anchor='NE',)
newax.imshow(logo)
newax.axis('off')
ax2.set_title('Filter Pressure Drop over Trace', loc='center')
ax2.plot(cycle_time, actual_speed, linewidth=0.9, alpha=0.75)
ax2.set_ylabel('Speed (Km/h)')
ax2.set_xlabel('Time (s)')
ax2.set_xlim([-5,1820])
ax2.set_xticks([0, 200, 400, 600, 800, 1000, 1200, 1400, 1600, 1800])
ax2.set_ylim([-0.5,140])
ax3 = ax2.twinx()
ax3.plot(cycle_time, filterPressure, color = 'red')
ax3.set_ylabel('Filter Pressure Drop (Pa)')

fig3, ax8 = plt.subplots(figsize=(16.53,11.69)) #CVS Flow Rate Vs Trace
plt.subplots_adjust(left=0.04, bottom=0.06, right=0.94, top=0.93, wspace=0.2, hspace=0.23)
newax = fig3.add_axes([0.84, 0.881, 0.124, 0.115], anchor='NE',)
newax.imshow(logo)
newax.axis('off')
ax8.set_title('CVS Flow Rate over Trace', loc='center')
ax8.plot(cycle_time, actual_speed, linewidth=0.9, alpha=0.75)
ax8.set_ylabel('Speed (Km/h)')
ax8.set_xlabel('Time (s)')
ax8.set_xlim([-5,1820])
ax8.set_xticks([0, 200, 400, 600, 800, 1000, 1200, 1400, 1600, 1800])
ax8.set_ylim([-0.5,140])
ax5 = ax8.twinx()
ax5.plot(cycle_time, cvsFlow, color = 'red',linewidth=0.5)
ax5.set_ylabel('CVS Flow Rate (m3/min)')

fig4, ax6 = plt.subplots(figsize=(16.53,11.69)) #SAO Flow Rate Vs Trace
plt.subplots_adjust(left=0.04, bottom=0.06, right=0.94, top=0.93, wspace=0.2, hspace=0.23)
newax = fig4.add_axes([0.84, 0.881, 0.124, 0.115], anchor='NE',)
newax.imshow(logo)
newax.axis('off')
ax6.set_title('SAO Flow Rate over Trace', loc='center')
ax6.plot(cycle_time, actual_speed, linewidth=0.9, alpha=0.75)
ax6.set_ylabel('Speed (Km/h)')
ax6.set_xlabel('Time (s)')
ax6.set_xlim([-5,1820])
ax6.set_xticks([0, 200, 400, 600, 800, 1000, 1200, 1400, 1600, 1800])
ax6.set_ylim([-0.5,140])
ax7 = ax6.twinx()
ax7.plot(cycle_time, exhaustFlow, color = 'red',linewidth=1)
ax7.set_ylabel('SAO Flow Rate (m3/s)')

fig5, ax8 = plt.subplots(figsize=(16.53,11.69)) #CO2 Vs Trace
plt.subplots_adjust(left=0.04, bottom=0.06, right=0.94, top=0.93, wspace=0.2, hspace=0.23)
newax = fig5.add_axes([0.84, 0.881, 0.124, 0.115], anchor='NE',)
newax.imshow(logo)
newax.axis('off')
ax8.set_title('CO2 Rate over trace', loc='center')
ax8.plot(cycle_time, actual_speed, linewidth=0.9, alpha=0.75)
ax8.set_ylabel('Speed (Km/h)')
ax8.set_xlabel('Time (s)')
ax8.set_xlim([-5,1820])
ax8.set_xticks([0, 200, 400, 600, 800, 1000, 1200, 1400, 1600, 1800])
ax8.set_ylim([-0.5,140])
diluteOnOff = diluteCO2Val[0]
tailpipeOnOff = postcatCO2Rate[0]
precatOnOff = precatCO2Val[0]
if type(diluteOnOff or tailpipeOnOff or precatOnOff ) == float:
     ax9 = ax8.twinx()
     ax9.plot(cycle_time, postcatCO2Rate, color = 'red',linewidth=1, label='Postcat Line')
     ax9.plot(cycle_time, diluteCO2Val, color = 'green',linewidth=1, label='Dilute Line')
     ax9.plot(cycle_time, precatCO2Val, color = 'orange',linewidth=1, label='Precat Line')
     ax9.set_ylabel('CO2 Rate (g/s)')  
     fig5.legend(ncol=3, loc='upper center')

fig6, ax10 = plt.subplots(figsize=(16.53,11.69)) #CO Vs Trace
plt.subplots_adjust(left=0.04, bottom=0.06, right=0.94, top=0.93, wspace=0.2, hspace=0.23)
newax = fig6.add_axes([0.84, 0.881, 0.124, 0.115], anchor='NE',)
newax.imshow(logo)
newax.axis('off')
ax10.set_title('CO Rate over trace', loc='center')
ax10.plot(cycle_time, actual_speed, linewidth=0.9, alpha=0.75)
ax10.set_ylabel('Speed (Km/h)')
ax10.set_xlabel('Time (s)')
ax10.set_xlim([-5,1820])
ax10.set_xticks([0, 200, 400, 600, 800, 1000, 1200, 1400, 1600, 1800])
ax10.set_ylim([-0.5,140])
diluteOnOff = diluteCORate[0]
tailpipeOnOff = postcatCORate[0]
precatOnOff = precatCORate[0]
if type(diluteOnOff or tailpipeOnOff or precatOnOff ) == float:
    ax11 = ax10.twinx()
    ax11.plot(cycle_time, postcatCORate, color = 'red',linewidth=1, label='Postcat Line')
    ax11.plot(cycle_time, diluteCORate, color = 'green',linewidth=1, label='Dilute Line')
    ax11.plot(cycle_time, precatCORate, color = 'orange',linewidth=1, label='Precat Line')
    ax11.set_ylabel('Postcat CO Rate (g/s)')
    fig6.legend(ncol=3, loc='upper center')

fig7, ax12 = plt.subplots(figsize=(16.53,11.69)) #THC Vs Trace
plt.subplots_adjust(left=0.04, bottom=0.06, right=0.94, top=0.93, wspace=0.2, hspace=0.23)
newax = fig7.add_axes([0.84, 0.881, 0.124, 0.115], anchor='NE',)
newax.imshow(logo)
newax.axis('off')
ax12.set_title('THC Rate over trace', loc='center')
ax12.plot(cycle_time, actual_speed, linewidth=0.9, alpha=0.75)
ax12.set_ylabel('Speed (Km/h)')
ax12.set_xlabel('Time (s)')
ax12.set_xlim([-5,1820])
ax12.set_xticks([0, 200, 400, 600, 800, 1000, 1200, 1400, 1600, 1800])
ax12.set_ylim([-0.5,140])
diluteOnOff = diluteTHCRate[0]
tailpipeOnOff = postcatTHCRate[0]
precatOnOff = precatTHCRate[0]
if type(diluteOnOff or tailpipeOnOff or precatOnOff ) == float:
    ax13 = ax12.twinx()
    ax13.plot(cycle_time, postcatTHCRate, color = 'red',linewidth=1, label='Postcat Line')
    ax13.plot(cycle_time, diluteTHCRate, color = 'green',linewidth=1, label='Dilute Line')
    ax13.plot(cycle_time, precatTHCRate, color = 'orange',linewidth=1, label='Precat Line')
    ax13.set_ylabel('Postcat THC Rate (g/s)')
    fig7.legend(ncol=3, loc='upper center')

fig8, ax14 = plt.subplots(figsize=(16.53,11.69)) #NOx Vs Trace
plt.subplots_adjust(left=0.04, bottom=0.06, right=0.94, top=0.93, wspace=0.2, hspace=0.23)
newax = fig8.add_axes([0.84, 0.881, 0.124, 0.115], anchor='NE',)
newax.imshow(logo)
newax.axis('off')
ax14.set_title('NOx Rate over trace', loc='center')
ax14.plot(cycle_time, actual_speed, linewidth=0.9, alpha=0.75)
ax14.set_ylabel('Speed (Km/h)')
ax14.set_xlabel('Time (s)')
ax14.set_xlim([-5,1820])
ax14.set_xticks([0, 200, 400, 600, 800, 1000, 1200, 1400, 1600, 1800])
ax14.set_ylim([-0.5,140])
diluteOnOff = diluteNOxRate[0]
tailpipeOnOff = postcatNOxRate[0]
precatOnOff = precatNOxRate[0]
if type(diluteOnOff or tailpipeOnOff or precatOnOff ) == float:
    ax15 = ax14.twinx()
    ax15.plot(cycle_time, postcatNOxRate, color = 'red',linewidth=1, label='Postcat Line')
    ax15.plot(cycle_time, diluteNOxRate, color = 'green',linewidth=1, label='Dilute Line')
    ax15.plot(cycle_time, precatNOxRate, color = 'orange',linewidth=1, label='Precat Line')
    ax15.set_ylabel('Postcat NOx Rate (g/s)')
    fig8.legend(ncol=3, loc='upper center')

fig9, ax16 = plt.subplots(figsize=(16.53,11.69)) #CH4 Vs Trace
plt.subplots_adjust(left=0.04, bottom=0.06, right=0.94, top=0.93, wspace=0.2, hspace=0.23)
newax = fig9.add_axes([0.84, 0.881, 0.124, 0.115], anchor='NE',)
newax.imshow(logo)
newax.axis('off')
ax16.set_title('CH4 Rate over trace', loc='center')
ax16.plot(cycle_time, actual_speed, linewidth=0.9, alpha=0.75)
ax16.set_ylabel('Speed (Km/h)')
ax16.set_xlabel('Time (s)')
ax16.set_xlim([-5,1820])
ax16.set_xticks([0, 200, 400, 600, 800, 1000, 1200, 1400, 1600, 1800])
ax16.set_ylim([-0.5,140])
diluteOnOff = diluteCH4Rate[0]
tailpipeOnOff = postcatCH4Rate[0]
precatOnOff = precatCH4Rate[0]
if type(diluteOnOff or tailpipeOnOff or precatOnOff ) == float:
    ax17 = ax16.twinx()
    ax17.plot(cycle_time, postcatCH4Rate, color = 'red',linewidth=1, label='Postcat Line')
    ax17.plot(cycle_time, diluteCH4Rate, color = 'green',linewidth=1, label='Dilute Line')
    ax17.plot(cycle_time, precatCH4Rate, color = 'orange',linewidth=1, label='Precat Line')
    ax17.set_ylabel('Postcat CH4 Rate (g/s)')
    fig9.legend(ncol=3, loc='upper center')

fig10, ax18 = plt.subplots(figsize=(16.53,11.69)) #Dyno Force Vs Trace
plt.subplots_adjust(left=0.04, bottom=0.06, right=0.94, top=0.93, wspace=0.2, hspace=0.23)
newax = fig10.add_axes([0.84, 0.881, 0.124, 0.115], anchor='NE',)
newax.imshow(logo)
newax.axis('off')
ax18.set_title('Dyno Force over trace', loc='center')
ax18.plot(cycle_time, actual_speed, linewidth=0.9, alpha=0.75)
ax18.set_ylabel('Speed (Km/h)')
ax18.set_xlabel('Time (s)')
ax18.set_xlim([-5,1820])
ax18.set_xticks([0, 200, 400, 600, 800, 1000, 1200, 1400, 1600, 1800])
ax18.set_ylim([-0.5,140])
ax19 = ax18.twinx()
ax19.plot(cycle_time, dynoForceData, color = 'red',linewidth=1)
ax19.set_ylabel('Dyno Force (N)')

fig11, ax20 = plt.subplots(figsize=(16.53,11.69)) #Dyno Power Vs Trace
plt.subplots_adjust(left=0.04, bottom=0.06, right=0.94, top=0.93, wspace=0.2, hspace=0.23)
newax = fig11.add_axes([0.84, 0.881, 0.124, 0.115], anchor='NE',)
newax.imshow(logo)
newax.axis('off')
ax20.set_title('Dyno Power over trace', loc='center')
ax20.plot(cycle_time, actual_speed, linewidth=0.9, alpha=0.75)
ax20.set_ylabel('Speed (Km/h)')
ax20.set_xlabel('Time (s)')
ax20.set_xlim([-5,1820])
ax20.set_xticks([0, 200, 400, 600, 800, 1000, 1200, 1400, 1600, 1800])
ax20.set_ylim([-0.5,140])
ax21 = ax20.twinx()
ax21.plot(cycle_time, dynoPowerData, color = 'red',linewidth=1)
ax21.set_ylabel('Dyno Power (mW)')

plt.rcParams["savefig.directory"] = os.chdir(os.path.dirname(folder_path_comp))
filename = PdfPages(filecut + " Analysis.pdf")
filename.savefig(fig)
filename.savefig(fig2)
filename.savefig(fig3)
filename.savefig(fig4)
filename.savefig(fig5)
filename.savefig(fig6)
filename.savefig(fig7)
filename.savefig(fig8)
filename.savefig(fig9)
filename.savefig(fig10)
filename.savefig(fig11)
filename.close()

print('Analysis Complete')