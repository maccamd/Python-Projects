#VDC Extra Analysis Package

from tkinter.constants import END
import openpyxl as xl
from openpyxl.utils import get_column_letter
import tkinter as tk
from tkinter import Image, filedialog
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages
from PIL import Image
import time

#----tkinter----#

root = tk.Tk()
root.withdraw()

#----Functions----#
start = time.time()

file_path = filedialog.askopenfilename() #taken from tkinter method
path = open(file_path, "rb") #rb is reading in binary
print(file_path)

wb = xl.load_workbook(path)

sheet = wb['Summary']   #inital data gathering, filename
filename = sheet['B1'].value 
filecut = filename[:-20] #takes out last 20 characters so project vehiclename and testid are left in theory

sheet = wb['ContinuousData']  #wb['Continous'] in unprocessed reports     #wb.active is sheet the file opens on

actspeed, DCclamp1, DCclamp2, DCclamp3, DCclamp4, FilterDrop, cvsflow = "", "", "", "", "", "", "" #Empty string that actual speed will go into

logo = Image.open('MPTLogo.png') # for plotting purposes

#loops to find column id in the continous data
for row_cells in sheet.iter_rows(min_row=1, max_row=1):
    for cell in row_cells:
        if 'DA Actual Speed' in cell.value or 'SpeedFeedback' in cell.value:
            actspeed = get_column_letter(cell.column)
            
for row_cells in sheet.iter_rows(min_row=1, max_row=1):
        for cell in row_cells:
            if 'Phase' in cell.value:
                phasenum = get_column_letter(cell.column)

for row_cells in sheet.iter_rows(min_row=1, max_row=1):
        for cell in row_cells:
            if 'Yokogawa wt5000REESSCurrent' in cell.value or 'Hioki 3390FundamentalIdc1' in cell.value:
                DCclamp1 = get_column_letter(cell.column)
                
for row_cells in sheet.iter_rows(min_row=1, max_row=1):
        for cell in row_cells:
            if 'Yokogawa wt5000REESSCurrent2' in cell.value  or 'Hioki 3390FundamentalIdc2' in cell.value:
                DCclamp2 = get_column_letter(cell.column)
                
for row_cells in sheet.iter_rows(min_row=1, max_row=1):
        for cell in row_cells:
            if 'Yokogawa wt5000REESSCurrent3' in cell.value  or 'Hioki 3390FundamentalIdc3' in cell.value:
                DCclamp3 = get_column_letter(cell.column) 

for row_cells in sheet.iter_rows(min_row=1, max_row=1):
        for cell in row_cells:
            if 'Yokogawa wt5000REESSCurrent4' in cell.value  or 'Hioki 3390FundamentalIdc4' in cell.value:
                DCclamp4 = get_column_letter(cell.column)

for row_cells in sheet.iter_rows(min_row=1, max_row=1):
        for cell in row_cells:
            if 'FilterPressureDrop' in cell.value:
                FilterDrop = get_column_letter(cell.column)
                
for row_cells in sheet.iter_rows(min_row=1, max_row=1):
        for cell in row_cells:
            if 'CVSFlowRate' in cell.value:
                cvsflow = get_column_letter(cell.column)
          
#gets data out of columns above and puts into variables
cycle_time = [sheet['B' + str(row)].value for row in range(3, sheet.max_row + 1)] #PlotTime Column G for RDEC 2 unprocessed reports
actual_speed = [sheet[actspeed + str(row)].value for row in range(3, sheet.max_row + 1)] #DA Actual Speed
phase = [sheet[phasenum + str(row)].value for row in range(3, sheet.max_row + 1)] #phase
current_clamp1 = [sheet[DCclamp1 + str(row)].value for row in range(3, sheet.max_row + 1)] #Current Clamp 1
current_clamp2 = [sheet[DCclamp2 + str(row)].value for row in range(3, sheet.max_row + 1)] #Current Clamp 2
current_clamp3 = [sheet[DCclamp3 + str(row)].value for row in range(3, sheet.max_row + 1)] #Current Clamp 3
current_clamp4 = [sheet[DCclamp4 + str(row)].value for row in range(3, sheet.max_row + 1)] #Current Clamp 4
FilterPressure = [sheet[FilterDrop + str(row)].value for row in range(3, sheet.max_row + 1)] #Filter Pressure Drop
CVSFlow = [sheet[cvsflow + str(row)].value for row in range(3, sheet.max_row + 1)] #CVS Flow Rate

#plotting
fig, ax = plt.subplots(figsize=(11.69,8.27))
plt.subplots_adjust(left=0.05, bottom=0.06, right=0.93, top=0.93, wspace=0.2, hspace=0.23)
newax = fig.add_axes([0.84, 0.881, 0.124, 0.115], anchor='NE',)
newax.imshow(logo)
newax.axis('off')
ax.set_title('REESS Balance over Trace', loc='center')
ax.plot(cycle_time, actual_speed)
ax.set_ylabel('Speed (Km/h)')
ax.set_xlabel('Time (s)')
ax.set_xlim([-5,1820])
ax.set_xticks([0, 200, 400, 600, 800, 1000, 1200, 1400, 1600, 1800])
ax.set_ylim([-0.5,140])
ax1 = ax.twinx()
ax1.axhline(y=0, color='g', linestyle='--')
ax1.plot(cycle_time, current_clamp1, color = 'red')
ax1.set_yticks([-600, -550, -500, -450, -400, -350, -300, -250, -200, -150, -100, -50, 0, 50, 100, 150, 200])
ax1.set_ylabel('Current (A)')
Clamp2OnOff = current_clamp2[0]
Clamp3OnOff = current_clamp3[0]
Clamp4OnOff = current_clamp4[0]

if Clamp2OnOff > 1:
    ax1.plot(cycle_time, current_clamp2, color = 'grey')  
if Clamp3OnOff > 1:
    ax1.plot(cycle_time, current_clamp3, color = 'orange')  
if Clamp4OnOff > 1:
    ax1.plot(cycle_time, current_clamp4, color = 'purple')          

fig2, ax2 = plt.subplots(figsize=(11.69,8.27))
plt.subplots_adjust(left=0.05, bottom=0.06, right=0.92, top=0.93, wspace=0.2, hspace=0.23)
newax = fig2.add_axes([0.84, 0.881, 0.124, 0.115], anchor='NE',)
newax.imshow(logo)
newax.axis('off')
ax2.set_title('Filter Pressure Drop over Trace', loc='center')
ax2.plot(cycle_time, actual_speed)
ax2.set_ylabel('Speed (Km/h)')
ax2.set_xlabel('Time (s)')
ax2.set_xlim([-5,1820])
ax2.set_xticks([0, 200, 400, 600, 800, 1000, 1200, 1400, 1600, 1800])
ax2.set_ylim([-0.5,140])
ax3 = ax2.twinx()
ax3.plot(cycle_time, FilterPressure, color = 'red')
ax3.set_ylabel('Filter Pressure Drop (Pa)')

fig3, ax4 = plt.subplots(figsize=(11.69,8.27))
plt.subplots_adjust(left=0.05, bottom=0.06, right=0.92, top=0.93, wspace=0.2, hspace=0.23)
newax = fig3.add_axes([0.84, 0.881, 0.124, 0.115], anchor='NE',)
newax.imshow(logo)
newax.axis('off')
ax4.set_title('CVS Flow Rate over Trace', loc='center')
ax4.plot(cycle_time, actual_speed)
ax4.set_ylabel('Speed (Km/h)')
ax4.set_xlabel('Time (s)')
ax4.set_xlim([-5,1820])
ax4.set_xticks([0, 200, 400, 600, 800, 1000, 1200, 1400, 1600, 1800])
ax4.set_ylim([-0.5,140])
ax5 = ax4.twinx()
ax5.plot(cycle_time, CVSFlow, color = 'red')
ax5.set_ylabel('CVS Flow Rate (m3/min)')
#plt.show()

filename = PdfPages(filecut + " Analysis.pdf")
filename.savefig(fig)
filename.savefig(fig2)
filename.savefig(fig3)
filename.close()

end = time.time()   
print(end-start) 